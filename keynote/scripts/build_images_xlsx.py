#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/user/collaboration-intelligence/keynote/deliverables/Bioadaptability_Image_Sources.xlsx"

# slide, imagen/concepto, autor, fuente, url, licencia, atribucion_requerida, estado
CANDIDATES = [
("15", "Redes de raices expuestas / metafora de micelio", "Aaron Escobar",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Exposed_mango_tree_roots.jpg",
 "CC BY 2.0", "\"Exposed mango tree roots\" by Aaron Escobar, CC BY 2.0, via Wikimedia Commons",
 "Candidata via busqueda — verificar pagina y version exacta de licencia antes de publicar (fetch directo bloqueado en este entorno)"),
("15", "Micelio de hongos micorrizicos arbusculares, microscopia falso color", "Ella Oyarte-Galvez (AMOLF)",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Mycelium_of_arbuscular_mycorrhizal_fungi_with_false_color.png",
 "CC BY-SA", "\"Mycelium of arbuscular mycorrhizal fungi with false color\" by Oyarte-Galvez/AMOLF, CC BY-SA, via Wikimedia Commons",
 "Candidata via busqueda — verificar version exacta de CC BY-SA antes de publicar"),
("15", "Micelio visible en el suelo del bosque", "PerytonMango",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Mycelium_in_forest_floor.jpg",
 "CC BY-SA 4.0", "\"Mycelium in forest floor\" by PerytonMango, CC BY-SA 4.0, via Wikimedia Commons",
 "Candidata via busqueda — verificar antes de publicar"),
("16", "Enjambre de abejas en racimo", "Mark Osgatharp",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Bee_Swarm.JPG",
 "CC BY-SA 3.0", "\"Bee Swarm\" by Mark Osgatharp, CC BY-SA 3.0, via Wikimedia Commons",
 "Candidata via busqueda — verificar antes de publicar"),
("17", "Murmuracion de estorninos", "Walter Baxter (via Geograph)",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Starling_murmuration.jpg",
 "CC BY-SA 2.0", "\"Starling murmuration\" by Walter Baxter, CC BY-SA 2.0, via Wikimedia Commons",
 "Candidata via busqueda — verificar antes de publicar"),
("17 (alt)", "Banco de peces / bola de cebo (defensa colectiva emergente)", "Kandu 1",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Moofushi_Kandu_fish.jpg",
 "CC BY-SA 2.5", "\"Moofushi Kandu fish\" by Kandu 1, CC BY-SA 2.5, via Wikimedia Commons",
 "Candidata via busqueda — verificar antes de publicar"),
("18", "Macrofagos en mitosis tras ingerir una celula fungica (inmunidad)", "Carolina Coelho",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Macrophages_undergo_mitosis_after_ingesting_a_fungal_cell.jpg",
 "CC BY-SA 4.0", "\"Macrophages undergo mitosis after ingesting a fungal cell\" by Carolina Coelho, CC BY-SA 4.0, via Wikimedia Commons",
 "Candidata via busqueda (finalista European Science Photo Competition 2015) — verificar antes de publicar"),
("19", "Bosque antiguo, dosel de copas (resiliencia/biodiversidad)", "Fæ (carga)",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Old_growth_coastal_forest.jpg",
 "Dominio publico", "\"Old growth coastal forest\", dominio publico, via Wikimedia Commons",
 "Candidata via busqueda — verificar antes de publicar"),
("19", "Regeneracion natural de bosques tras incendio (categoria, buscar archivo especifico)", "Multiples autores",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/Category:Natural_regeneration_of_forests",
 "Variable segun archivo", "Ver archivo especifico dentro de la categoria y su licencia individual",
 "Enlace a CATEGORIA, no a un archivo — elegir imagen especifica dentro y verificar licencia antes de publicar"),
("19 (alt)", "Dosel de arboles, Alas Purwo (Wiki Loves Earth 2020)", "Aphrodity13",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Tree_Canopy.jpg",
 "CC BY-SA 4.0", "\"Tree Canopy\" by Aphrodity13, CC BY-SA 4.0, via Wikimedia Commons",
 "Candidata via busqueda — verificar antes de publicar"),
("bonus/16", "Bivac de hormigas legionarias (estructura viva autoorganizada)", "Geoff Gallice",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Army_ant_bivouac.jpg",
 "CC BY 2.0", "\"Army ant bivouac\" by Geoff Gallice, CC BY 2.0, via Wikimedia Commons",
 "Candidata via busqueda — verificar antes de publicar"),
("bonus/16", "Hormigas cortadoras de hojas transportando carga (asignacion de tareas)", "Bandwagonman",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Leafcutter_ants_transporting_leaves.jpg",
 "CC BY-SA (multi)", "\"Leafcutter ants transporting leaves\" by Bandwagonman, CC BY-SA, via Wikimedia Commons",
 "Candidata via busqueda — verificar antes de publicar"),
("bonus (Acto IV, apertura)", "Red del Physarum polycephalum (moho mucilaginoso, diseno de red)", "Adaptado de Nakagaki et al.",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Physarum_polycephalum_network.jpg",
 "CC BY-SA 4.0", "\"Physarum polycephalum network\" adaptado de Nakagaki et al., CC BY-SA 4.0, via Wikimedia Commons",
 "Candidata via busqueda — verificar antes de publicar"),
("19 / cierre alternativo", "Arrecife de coral, Mar Rojo (biodiversidad/resiliencia)", "Hagainativ",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Red_sea_coral_reef.jpg",
 "CC BY-SA 3.0", "\"Red sea coral reef\" by Hagainativ, CC BY-SA 3.0, via Wikimedia Commons",
 "Candidata via busqueda — verificar antes de publicar"),
("19 / cierre alternativo", "Fotografia submarina de arrecife de coral", "Fæ (carga, fuente Public-domain-image.com)",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Underwater_photo_of_coral_reef.jpg",
 "Dominio publico", "\"Underwater photo of coral reef\", dominio publico, via Wikimedia Commons",
 "Candidata via busqueda — verificar antes de publicar"),
("apertura (Acto IV) / conceptual", "Pilares de la Creacion, Nebulosa del Aguila (Hubble, 2014, alta resolucion)", "NASA, ESA y el Hubble Heritage Team (STScI/AURA)",
 "Wikimedia Commons / NASA", "https://commons.wikimedia.org/wiki/File:Pillars_of_creation_2014_HST_WFC3-UVIS_full-res.jpg",
 "Dominio publico (NASA); credito ESA habitual", "\"NASA, ESA and the Hubble Heritage Team (STScI/AURA)\", dominio publico, via Wikimedia Commons",
 "Candidata via busqueda — imagen de NASA, dominio publico por politica estandar de NASA; verificar pagina antes de publicar"),
("11", "Mesa de sala de juntas vacia", "Subido a Wikimedia Commons (Unsplash)",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/File:Conference_room_table_(Unsplash).jpg",
 "Verificar en pagina (via Unsplash)", "Verificar texto de licencia exacto en la pagina antes de publicar",
 "Candidata nueva via busqueda — verificar licencia exacta antes de publicar"),
("35", "Arbol solitario en campo abierto al amanecer", "Multiples fotografos (ver Category:Solitary trees / Photo challenge 2019)",
 "Wikimedia Commons", "https://commons.wikimedia.org/wiki/Category:Solitary_trees",
 "Variable segun archivo", "Elegir imagen especifica dentro de la categoria y su licencia individual",
 "Enlace a CATEGORIA — elegir un archivo especifico (ej. Lone_tree_in_front_of_Dune_45_Sossusvlei_Namibia.jpg) y verificar licencia antes de publicar"),
]

# Slides needing a fresh sourcing pass (no candidate found/verified yet) — briefs only, no invented URLs
BRIEFS = [
("1", "Titulo — gradiente de amanecer sobre textura topografica/curvas de nivel", "abstract topographic contour lines mountain gradient dawn light",
 "Diseno abstracto propio (no requiere foto de stock); si se prefiere foto real, buscar en Unsplash/Pexels 'topographic aerial mountain ridge light'"),
("2", "Hoja seca cayendo, bosque estatico y nitido", "single falling leaf forest floor sharp focus stillness",
 "Wikimedia Commons / Unsplash — buscar 'falling leaf motion forest' con licencia CC clara"),
("3", "Mismo bosque en movimiento continuo (lluvia/viento, sin punto fijo)", "forest rain wind motion blur long exposure",
 "Wikimedia Commons / Unsplash — buscar 'forest rain long exposure motion blur'"),
("4", "Instrumento antiguo de medicion (sextante/veleta) agrietado sobre superficie moderna", "antique sextant weathervane cracked obsolete instrument",
 "Wikimedia Commons — categoria 'Sextants' o 'Weathervanes', elegir archivo con licencia CC clara"),
("5", "Bocetos vs. prototipo funcional, brecha comprimida", "hand sketch notebook vs working app prototype split screen",
 "Foto compuesta a disenar (no existe una unica foto de stock adecuada); tratamiento grafico propio recomendado"),
("6", "Escritorio simple, una laptop, luz dura direccional", "minimal desk single laptop harsh directional light solo founder",
 "Unsplash/Pexels — buscar 'minimal desk single laptop dramatic light', evitar cliches de 'startup hustle'"),
("7,8,9", "Fondo negro con tipografia — palabra/pregunta sola", "n/a — tratamiento tipografico", "Diseno propio, no requiere fotografia"),
("9 (alt)", "Un solo haz de luz en espacio arquitectonico vacio (escalera, silo)", "single light shaft empty stairwell silo interior architecture",
 "Wikimedia Commons — categoria 'Staircases' o 'Grain silos interior', elegir archivo con licencia CC clara"),
("10", "Plano/dibujo arquitectonico de un puente en construccion, siluetas pequenas", "bridge under construction blueprint workers silhouette scale",
 "Wikimedia Commons — categoria 'Bridges under construction', elegir archivo con licencia CC clara"),
("12", "Engranaje/gobernador mecanico industrial, aislado", "antique mechanical governor gear industrial isolated",
 "Wikimedia Commons — categoria 'Centrifugal governors' o 'Gears', elegir archivo con licencia CC clara"),
("13", "Estromatolito fosilizado / estratos geologicos en corte transversal", "stromatolite fossil cross section geological strata",
 "Wikimedia Commons — categoria 'Stromatolites', elegir archivo con licencia CC clara"),
("14", "Macro extrema de un solo hilo de hifa fungica", "hyphae macro extreme close-up fungal thread forest litter",
 "Wikimedia Commons — buscar 'hyphae macro photography', elegir archivo con licencia CC clara"),
("20", "Circulos superpuestos estilo pizarra, con marcas de borrado visibles", "hand-drawn overlapping circles chalkboard eraser marks revision",
 "Tratamiento grafico propio recomendado (no foto de stock)"),
("21", "Silla vacia a la cabecera de una mesa larga, resto desenfocado", "empty chair head of long table blurred background",
 "Wikimedia Commons/Unsplash — buscar 'empty chair head of table', elegir archivo con licencia CC clara"),
("22", "Piedras de rio lisas vs roca fracturada rugosa, misma materia", "smooth river stones vs jagged broken rock texture comparison",
 "Wikimedia Commons — categorias 'River stones' y 'Broken rock', elegir archivos con licencia CC clara"),
("23", "Boceto tecnico a mano, textura de papel de fax, pinchado en pared de fabrica", "hand-drawn technical blueprint sketch fax paper texture factory wall",
 "Recreacion grafica propia recomendada (imagen documental especifica del incendio de Aisin no esta disponible con licencia clara)"),
("24", "Hilo continuo cosido a traves de tela rasgada, remiendo visible", "single thread stitched torn fabric visible mend",
 "Wikimedia Commons/Unsplash — buscar 'visible mend stitched fabric', elegir archivo con licencia CC clara"),
("25", "Silla vacia comun en mesa redonda, sin robot ni brillo", "empty ordinary chair round table no tech no glow",
 "Wikimedia Commons/Unsplash — buscar 'empty chair round table minimal', elegir archivo con licencia CC clara"),
("26,27", "Auriculares de servicio al cliente sobre escritorio vacio", "customer service headset empty desk unused minimal",
 "Unsplash/Pexels — buscar 'headset desk minimal unused', elegir archivo con licencia CC clara"),
("28", "Silueta de una figura, luego repetida identica en fila, desvaneciendose", "silhouette figure repeated row identical fading edge frame",
 "Tratamiento grafico propio recomendado (efecto de repeticion dificil de encontrar como foto unica con licencia clara)"),
("29", "Muchas lamparas de escritorio encendidas en oficina oscura, sin tocarse", "many desk lamps lit dark open office floor separate pools of light",
 "Wikimedia Commons/Unsplash — buscar 'desk lamps dark office night', elegir archivo con licencia CC clara"),
("30", "Formas organicas entrelazadas tipo rompecabezas (silueta de plegado de proteina)", "abstract interlocking organic puzzle shapes protein fold silhouette",
 "Tratamiento grafico/generativo propio recomendado (no biologia literal)"),
("31", "Dos circulos de luz superpuestos en piso oscuro, uno reposicionandose", "two overlapping light circles dark floor repositioning",
 "Tratamiento grafico propio recomendado"),
("32", "Ciclo renderizado como espiral natural (corte de nautilus o anillos de arbol)", "nautilus shell cross section spiral OR tree rings cross section",
 "Wikimedia Commons — categoria 'Nautilus shells' o 'Tree rings', elegir archivo con licencia CC clara"),
("33", "Cordon o cuerda de tiro colgando en espacio industrial, retroiluminado", "pull cord rope hanging industrial space backlit isolated",
 "Wikimedia Commons/Unsplash — buscar 'pull cord industrial rope backlit', elegir archivo con licencia CC clara"),
("34", "Cuerda deshilachada, cortada limpiamente y reempalmada, reparacion visible", "frayed rope cleanly spliced repair visible",
 "Wikimedia Commons/Unsplash — buscar 'spliced rope repair visible', elegir archivo con licencia CC clara"),
]

def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Imagenes Candidatas"

    title_font = Font(name="Arial", size=14, bold=True, color="1F2937")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="B08D57")
    body_font = Font(name="Arial", size=9.5)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Bioadaptability — Fuentes de Imagenes"
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")
    note = ("IMPORTANTE: en esta sesion de ejecucion, el acceso de red saliente a commons.wikimedia.org, "
            "images.nasa.gov, unsplash.com y en.wikipedia.org estuvo bloqueado por politica organizacional "
            "del entorno (confirmado via el proxy de egress; no es un problema tecnico solucionable desde aqui). "
            "Esto significa que NO fue posible descargar los archivos de imagen ni cargar sus paginas directamente "
            "para confirmar visualmente el texto exacto de licencia. Cada URL de abajo fue localizada mediante "
            "busqueda web real (no inventada), pero debe verificarse abriendo la pagina de Wikimedia Commons "
            "directamente antes de usar la imagen en una presentacion publica.")
    ws["A2"] = note
    ws["A2"].font = Font(name="Arial", size=8.5, italic=True, color="B91C1C")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 60

    headers = ["slide", "imagen_concepto", "autor_fotografo", "fuente", "url", "licencia", "atribucion_requerida", "estado_verificacion"]
    header_row = 4
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    r = header_row + 1
    for row in CANDIDATES:
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = body_font; c.alignment = wrap; c.border = border
        ws.row_dimensions[r].height = 50
        r += 1

    widths = [14, 34, 26, 16, 46, 16, 46, 42]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:H{r-1}"

    # Sheet 2: briefs for slides without a verified candidate
    ws2 = wb.create_sheet("Briefs de Busqueda Pendiente")
    ws2["A1"] = "Slides que requieren sourcing de imagen antes de produccion final"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:D1")
    ws2["A2"] = ("Estas imagenes no pudieron confirmarse con una URL real y licencia verificada dentro de esta sesion "
                 "(bloqueo de red descrito arriba). Se entrega el brief creativo/de busqueda para que el equipo de diseno "
                 "las busque directamente en Wikimedia Commons, NASA image libraries, Unsplash o Pexels, o las produzca "
                 "como tratamiento grafico propio donde se indica.")
    ws2["A2"].font = Font(name="Arial", size=8.5, italic=True, color="6B7280")
    ws2["A2"].alignment = Alignment(wrap_text=True)
    ws2.merge_cells("A2:D2")
    ws2.row_dimensions[2].height = 45

    headers2 = ["slide", "concepto_visual", "terminos_de_busqueda_sugeridos", "recomendacion"]
    hr2 = 4
    for col, h in enumerate(headers2, start=1):
        c = ws2.cell(row=hr2, column=col, value=h)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    r2 = hr2 + 1
    for row in BRIEFS:
        for col, val in enumerate(row, start=1):
            c = ws2.cell(row=r2, column=col, value=val)
            c.font = body_font; c.alignment = wrap; c.border = border
        ws2.row_dimensions[r2].height = 46
        r2 += 1
    widths2 = [22, 40, 46, 50]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A5"
    ws2.auto_filter.ref = f"A{hr2}:D{r2-1}"

    # Sheet 3: licensing guardrails
    ws3 = wb.create_sheet("Guia de Licencias")
    ws3["A1"] = "Guia rapida antes de publicar cualquier imagen en la keynote"
    ws3["A1"].font = title_font
    ws3.merge_cells("A1:B1")
    guardrails = [
        ("Prioridad de fuentes", "Wikimedia Commons > NASA (images.nasa.gov, dominio publico por defecto) > Unsplash/Pexels (verificar licencia de cada foto individual) > instituciones publicas/museos."),
        ("Antes de usar cualquier imagen", "Abrir la pagina File: en Wikimedia Commons y confirmar: (1) el nombre exacto del autor, (2) la version exacta de la licencia (ej. CC BY-SA 4.0 vs 2.0 no son intercambiables), (3) si requiere atribucion visible en el slide o basta con el documento de fuentes."),
        ("Nunca usar si...", "La licencia no esta clara, el archivo dice 'fair use' o 'all rights reserved', o no se puede confirmar el autor original (algunas paginas de Commons redistribuyen contenido de terceros sin licencia clara)."),
        ("NASA", "Las imagenes producidas por la NASA son tipicamente de dominio publico en EE.UU. (17 U.S.C. § 105), pero verificar si la imagen incluye contribuciones de la ESA u otra agencia, que pueden requerir credito adicional."),
        ("Formato de atribucion en notas del orador o creditos finales", "\"[Titulo]\" por [Autor], [Licencia], via [Fuente] — [URL]"),
        ("Limitacion de esta entrega", "El acceso de red a los repositorios de imagenes estuvo bloqueado en el entorno de ejecucion de esta sesion; todas las URLs listadas se encontraron via busqueda real pero no se verificaron por carga directa de pagina. Verificar cada una antes de la produccion final del PPTX con imagenes reales."),
    ]
    r3 = 3
    for label, text in guardrails:
        ws3.cell(row=r3, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws3.cell(row=r3, column=1).alignment = wrap
        ws3.cell(row=r3, column=2, value=text).font = body_font
        ws3.cell(row=r3, column=2).alignment = wrap
        ws3.row_dimensions[r3].height = 55
        r3 += 1
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 80

    wb.save(OUT)
    print("Saved", OUT, "candidates:", len(CANDIDATES), "briefs:", len(BRIEFS))

if __name__ == "__main__":
    build()
