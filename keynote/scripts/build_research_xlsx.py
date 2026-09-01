#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/user/collaboration-intelligence/keynote/deliverables/Bioadaptability_Research.xlsx"

HEADERS = ["autor", "organizacion", "fuente", "url", "anio", "idea_central",
           "evidencia", "cita", "categoria", "score_1_5", "relevancia_tesis",
           "utilizado_en_keynote", "slide_relacionada"]

# (autor, organizacion, fuente/titulo, url, anio, idea, evidencia, cita, categoria, score, relevancia, uso, slide)
ROWS = [
("Anita Woolley, Christopher Chabris, Alex Pentland, Nada Hashmi, Thomas Malone", "Carnegie Mellon / MIT",
 "Evidence for a Collective Intelligence Factor in the Performance of Human Groups", "https://www.science.org/doi/10.1126/science.1193147",
 2010, "Existe un factor estadistico 'c' de inteligencia colectiva que predice el desempeno grupal en tareas diversas.",
 "Estudio experimental con grupos; c correlaciona debilmente con el CI promedio/maximo del grupo, fuertemente con sensibilidad social promedio y equidad en los turnos de habla.",
 "c is not strongly correlated with the average or maximum individual intelligence of group members but is correlated with the average social sensitivity of group members.",
 "collective-intelligence", 5, "Pilar central: la inteligencia de un grupo no es la suma de CI individuales.", "Si", "20"),

("Anita Woolley, Thomas Malone", "Harvard Business Review",
 "Defend Your Research: What Makes a Team Smarter? More Women", "https://hbr.org/2011/06/defend-your-research-what-makes-a-team-smarter-more-women",
 2011, "Divulga el hallazgo del factor c a audiencia de management.",
 "Popularizacion del estudio de 2010 para lideres empresariales.",
 "There's little correlation between a group's collective intelligence and the IQ scores of its individual members.",
 "collective-intelligence", 3, "Refuerza el hallazgo central de forma citable en vivo.", "Si", "20"),

("Marcus Crede, Peter Howardson", "Journal of Applied Psychology",
 "The Structure of Group Task Performance: A Second Look at the Evidence for the 'Collective Intelligence' Construct", "https://www.researchgate.net/publication/308399872",
 2017, "Cuestiona metodologicamente la existencia de un factor c general.",
 "Reanalisis de los datos originales; argumenta que no sostienen un factor general unico.",
 "[The data do] not support the inference that a general factor can explain substantial variation in performance across a wide cross section of group tasks.",
 "contradiction", 4, "Evita que la tesis se presente como ciencia cerrada; honestidad intelectual.", "Si", "20"),

("Anita Woolley, Yeonjeong Kim, Thomas Malone", "SSRN",
 "Measuring Collective Intelligence in Groups: A Reply to Crede and Howardson", "https://www.ssrn.com/abstract=3187373",
 2018, "Replica de Woolley/Malone a la critica metodologica.",
 "Defiende el factor c frente a la reanalisis de Crede & Howardson.",
 "", "collective-intelligence", 2, "Muestra que el debate academico sigue abierto.", "Mencion en notas", "20"),

("Young Ji Kim, Anita Woolley et al.", "Cognitive Research: Principles and Implications",
 "g versus c: comparing individual and collective intelligence across two meta-analyses", "https://link.springer.com/article/10.1186/s41235-021-00285-2",
 2021, "Meta-analisis que intenta zanjar el debate del factor c.",
 "Replicaciones en configuraciones presenciales, online e hibridas con resultados mixtos.",
 "", "contradiction", 3, "Refuerza que el debate cientifico no esta resuelto.", "Mencion en notas", "20"),

("Thomas W. Malone", "MIT Center for Collective Intelligence",
 "Superminds: The Surprising Power of People and Computers Thinking Together (libro)", "https://www.hachettebookgroup.com/titles/thomas-w-malone/superminds/9780316349109/",
 2018, "Jerarquias, mercados, democracias y comunidades son todos 'superminds'; la IA se convierte en un nuevo miembro de esos superminds.",
 "Marco teorico integrador sobre sistemas humano-maquina de pensamiento colectivo.",
 "", "collective-intelligence", 4, "Sustenta el Acto VII: de inteligencia individual a inteligencia de sistema.", "Si", "29,30"),

("Scott E. Page", "Princeton University Press",
 "The Difference: How the Power of Diversity Creates Better Groups, Firms, Schools, and Societies (libro)", "https://press.princeton.edu/books/hardcover/9780691176888/the-diversity-bonus",
 2007, "'La diversidad vence a la habilidad' en problemas dificiles sin un solucionador dominante, con buena comunicacion y objetivos compartidos.",
 "Modelos formales y estudios de caso sobre resolucion de problemas en grupos heterogeneos.",
 "Diversity trumps ability.",
 "collective-intelligence", 4, "Base del principio de diversidad, con matiz de condiciones.", "Si", "22"),

("Scott E. Page", "Wharton School",
 "Diversity Paradoxes (working paper)", "https://faculty.wharton.upenn.edu/wp-content/uploads/2015/04/Diversity_Paradoxes_1.pdf",
 2015, "El propio Page mapea cuando el 'bono de diversidad' falla o se revierte.",
 "Analisis de condiciones limite del efecto de diversidad.",
 "", "contradiction", 4, "Contrapeso necesario a la slide de diversidad.", "Si", "22"),

("James Surowiecki", "Doubleday",
 "The Wisdom of Crowds (libro)", "https://en.wikipedia.org/wiki/The_Wisdom_of_Crowds",
 2004, "Los juicios agregados de un grupo superan a los expertos bajo 4 condiciones: diversidad de opinion, independencia, descentralizacion y un mecanismo de agregacion.",
 "El propio libro dedica capitulos a cuando las multitudes fallan (manadas, comites, burbujas).",
 "", "collective-intelligence", 3, "Las 4 condiciones son el matiz que las versiones populares omiten.", "Mencion en notas", "17,20"),

("Duncan J. Watts", "Crown Business",
 "Everything Is Obvious: How Common Sense Fails Us (libro)", "https://en.wikipedia.org/wiki/Everything_Is_Obvious",
 2011, "El razonamiento de sentido comun sobre sistemas sociales es sistematicamente poco fiable.",
 "Argumenta que los efectos de 'sabiduria de las multitudes' son mas fragiles/condicionales de lo que se cree.",
 "", "contradiction", 3, "Refuerza la necesidad de evidencia, no intuicion, para decisiones colectivas.", "Mencion en notas", "20"),

("Matthew Salganik, Peter Dodds, Duncan Watts", "Science",
 "Experimental Study of Inequality and Unpredictability in an Artificial Cultural Market", "https://www.science.org/doi/10.1126/science.1121066",
 2006, "Cuando los participantes ven las elecciones de otros, la influencia social genera desigualdad y impredecibilidad en lo que se vuelve 'popular'.",
 "Experimento de mercado musical artificial con y sin visibilidad social.",
 "The best songs rarely did poorly, and the worst rarely did well, but almost any other result was possible.",
 "contradiction", 4, "Anticipa el mecanismo detras del riesgo de homogenizacion con IA.", "Si", "28",),

("Jan Lorenz, Heiko Rauhut, Frank Schweitzer, Dirk Helbing", "PNAS",
 "How Social Influence Can Undermine the Wisdom of Crowd Effect", "https://www.pnas.org/doi/10.1073/pnas.1008636108",
 2011, "Incluso una influencia social leve (ver las estimaciones de otros) reduce la diversidad de juicios independientes y destruye el efecto de sabiduria de las multitudes.",
 "Experimento controlado de estimacion numerica con y sin visibilidad de respuestas ajenas.",
 "", "contradiction", 5, "Mecanismo clave: hablar entre si puede volver al grupo menos sabio, no mas.", "Si", "22,28"),

("Donella Meadows", "Sustainability Institute",
 "Leverage Points: Places to Intervene in a System", "https://donellameadows.org/wp-content/userfiles/Leverage_Points.pdf",
 1999, "12 lugares para intervenir en un sistema; el de mayor apalancamiento es el paradigma/mentalidad sobre el que se construye el sistema.",
 "Marco de sistemas complejos aplicado a intervencion, ofrecido 'con humildad', no como formula mecanica.",
 "", "complexity", 4, "Sustenta que cambiar la organizacion requiere cambiar el paradigma, no solo el proceso.", "Mencion en notas", "32"),

("Dave Snowden, Mary Boone", "Harvard Business Review",
 "A Leader's Framework for Decision Making", "https://www.harvardbusiness.org/insight/navigating-complexity-a-new-map-for-a-new-territory/",
 2007, "Cynefin: los dominios claro, complicado, complejo y caotico requieren enfoques de decision distintos; en lo complejo, causa-efecto solo se ve en retrospectiva.",
 "Marco de sensemaking adoptado ampliamente en estrategia y liderazgo.",
 "", "complexity", 5, "Explica por que planificar no basta quo problemas complejos: hay que 'probar-sentir-responder'.", "Si", "10"),

("Eric Beinhocker", "Harvard Business School Press",
 "The Origin of Wealth: Evolution, Complexity, and the Radical Remaking of Economics (libro)", "https://www.harvard.com/book/origin_of_wealth_evolution_complexity_and_the_radical_remaking_of_economics/",
 2006, "La economia es un sistema adaptativo complejo y evolutivo (diferenciar-seleccionar-amplificar), no un sistema de equilibrio.",
 "Sintesis de economia de la complejidad.",
 "", "complexity", 3, "Cuestiona la nocion de 'ejecucion optima' como objetivo coherente.", "Mencion en notas", "10"),

("Geoffrey West", "Penguin Press",
 "Scale: The Universal Laws of Growth, Innovation, Sustainability, and the Pace of Life... (libro)", "https://www.penguinrandomhouse.com/books/314049/scale-by-geoffrey-west/",
 2017, "Las ciudades escalan superlinealmente y son abiertas; las empresas escalan sublinealmente y mueren de forma fiable.",
 "Analisis cuantitativo de escalamiento en ciudades, organismos y empresas.",
 "", "complexity", 4, "La estructura, no solo la capacidad, determina si el crecimiento se compone o decae.", "Mencion en notas", "19"),

("Nassim Nicholas Taleb", "Random House",
 "Antifragile: Things That Gain from Disorder (libro)", "https://en.wikipedia.org/wiki/Antifragile_(book)",
 2012, "Distingue sistemas fragiles, robustos y antifragiles (que mejoran con el estres/volatilidad).",
 "Marco conceptual sobre riesgo y adaptacion.",
 "", "complexity", 3, "La ejecucion eficiente puede ser fragil ante shocks; relevante para Acto IV/VIII.", "Mencion en notas", "19"),

("Ronald Heifetz", "Belknap/Harvard University Press",
 "Leadership Without Easy Answers (libro)", "https://umbrex.com/resources/frameworks/organization-frameworks/adaptive-leadership-framework-heifetz/",
 1994, "Distingue problemas tecnicos (solucion conocida) de retos adaptativos (sin manual, exigen cambiar valores/habitos).",
 "Marco de liderazgo adaptativo ampliamente citado.",
 "The single biggest failure of leadership is to treat adaptive challenges like technical problems.",
 "complexity", 5, "Explica por que ejecutar mas rapido no resuelve retos adaptativos.", "Si", "10,12"),

("Margaret Wheatley", "Berrett-Koehler",
 "Leadership and the New Science: Discovering Order in a Chaotic World (libro)", "https://margaretwheatley.com/books/leadership-and-the-new-science/",
 1992, "Aplica teoria del caos y autoorganizacion al liderazgo; las relaciones y el flujo de informacion generan orden, no el control.",
 "Sintesis entre fisica/sistemas complejos y teoria organizacional.",
 "", "complexity", 3, "Sustenta el cambio de liderazgo: de control a condiciones para que emerja el orden.", "Mencion en notas", "32"),

("Peter Senge", "Doubleday",
 "The Fifth Discipline: The Art and Practice of the Learning Organization (libro)", "https://en.wikipedia.org/wiki/The_Fifth_Discipline",
 1990, "El pensamiento sistemico integra maestria personal, modelos mentales, vision compartida y aprendizaje en equipo.",
 "Marco fundacional de las organizaciones que aprenden.",
 "In the long run, the only sustainable competitive advantage is your organization's ability to learn faster than the competition.",
 "org-learning", 4, "Frase ancla para el cierre: aprender juntos mas rapido que el entorno cambia.", "Si", "35"),

("Amy Edmondson", "Administrative Science Quarterly",
 "Psychological Safety and Learning Behavior in Work Teams", "https://web.mit.edu/curhan/www/docs/Articles/15341_Readings/Group_Performance/Edmondson%20Psychological%20safety.pdf",
 1999, "La seguridad psicologica del equipo (no la autoeficacia) predice el comportamiento de aprendizaje, que media el desempeno.",
 "Estudio de campo con 51 equipos.",
 "", "org-learning", 5, "Base empirica de por que la seguridad psicologica es la 'red' que permite circular la informacion.", "Si", "24"),

("Amy Edmondson", "Wiley",
 "The Fearless Organization: Creating Psychological Safety in the Workplace (libro)", "https://www.wiley.com/en-us/The+Fearless+Organization",
 2018, "Extiende el constructo de 1999 a un marco de liderazgo practico.",
 "Sintesis aplicada, incluye referencias a Project Aristotle de Google.",
 "", "org-learning", 3, "Vinculo directo con Project Aristotle de Google.", "Mencion en notas", "24"),

("Google re:Work (Project Aristotle)", "Google",
 "Project Aristotle", "https://fearlessorganizationscan.com/project-aristotle-explained",
 2015, "Analisis de 180+ equipos de Google: la seguridad psicologica importa mas que quien esta en el equipo.",
 "Estudio interno de gran escala que corrobora a Edmondson.",
 "", "org-learning", 4, "Validacion empresarial del hallazgo academico.", "Si", "24"),

("Ethan Mollick", "Portfolio/Penguin",
 "Co-Intelligence: Living and Working with AI (libro)", "https://www.forbes.com/sites/peterhigh/2024/05/07/ethan-mollick-on-the-four-rules-of-co-intelligence-with-ai/",
 2024, "Acuna la 'frontera irregular' (jagged frontier): el desempeno de la IA es impredeciblemente desigual entre tareas; el cuello de botella practico es saber donde esta la frontera.",
 "Sintesis de investigacion propia y de terceros sobre uso de IA en el trabajo del conocimiento.",
 "A landscape of sharp peaks where AI excels beyond human ability and deep valleys where it fails at tasks a child could master.",
 "AI-execution", 5, "Concepto central del Acto VI: la IA no reduce la necesidad de juicio, la desplaza.", "Si", "25,26,27"),

("Fabrizio Dell'Acqua, Edward McFowland III, Ethan Mollick, et al.", "HBS / BCG / Organization Science",
 "Navigating the Jagged Technological Frontier: Field Experimental Evidence...", "https://papers.ssrn.com/sol3/papers.cfm?abstract_id=4573321",
 2023, "758 consultores de BCG asignados aleatoriamente a condiciones sin IA / con GPT-4 / con GPT-4+entrenamiento en 18 tareas realistas.",
 "Dentro de la frontera de competencia de la IA: velocidad +25%, calidad +40%, finalizacion +12%; fuera de la frontera, sobreconfianza y peor desempeno.",
 "", "AI-execution", 5, "Evidencia experimental directa y citable de la 'ejecucion abundante'.", "Si", "5"),

("Sida Peng, Eirini Kalliamvakou, Peter Cihon, Mert Demirer", "GitHub / Microsoft / MIT (arXiv)",
 "The Impact of AI on Developer Productivity: Evidence from GitHub Copilot", "https://arxiv.org/abs/2302.06590",
 2023, "Ensayo controlado aleatorizado: desarrolladores con GitHub Copilot completaron una tarea estandarizada 55.8% mas rapido.",
 "Experimento aleatorizado (no un caso de estudio de proveedor); mayor beneficio en desarrolladores menos experimentados.",
 "", "AI-execution", 5, "Prueba cuantitativa de que ejecutar se vuelve radicalmente mas barato con IA.", "Si", "5"),

("Satya Nadella (comentario/analisis)", "Microsoft",
 "'Token capital vs. taste': cuando ejecutar es barato, el juicio vale la prima", "https://eliteagent.com/when-execution-is-cheap-judgment-is-worth-the-premium/",
 2025, "La IA escala el 'capital de tokens' (ejecucion); las empresas deben organizarse en torno al 'gusto/juicio' humano escaso.",
 "Analisis secundario que sintetiza declaraciones publicas de Nadella.",
 "Token capital scales the execution, while taste decides what to execute on.",
 "AI-execution", 4, "Formulacion ejecutiva casi identica a la tesis central del keynote.", "Si", "9,29"),

("Sam Altman / Greg Brockman", "OpenAI (via Fortune)",
 "Have good taste? It may just get you a job during the AI jobs apocalypse", "https://fortune.com/2026/02/27/openai-sam-altman-taste-get-jobseekers-hired-ai-jobpocalypse",
 2026, "OpenAI contrata cada vez mas por 'contexto, gusto y sentido de hacia donde va el campo', no solo habilidad tecnica ejecutable.",
 "Declaraciones reportadas de liderazgo de OpenAI sobre criterios de contratacion.",
 "We believe the best research teams are built through context, taste and a real feel for where the field is headed next.",
 "AI-execution", 4, "Evidencia de que incluso el lider de IA generativa valora el juicio sobre la ejecucion.", "Si", "9"),

("Jensen Huang", "NVIDIA (via CNBC / World Government Summit)",
 "Everyone is a programmer with generative AI", "https://www.cnbc.com/2023/05/30/everyone-is-a-programmer-with-generative-ai-nvidia-ceo-.html",
 2023, "La IA generativa colapsa la barrera tecnica para ejecutar codigo; el lenguaje de programacion pasa a ser el humano.",
 "Declaraciones publicas + analisis posterior sobre el 'AI talent thesis' de Huang.",
 "AI lowers the technical barrier, but it does not lower the thinking barrier. In many ways, it raises it.",
 "AI-execution", 4, "Cita directa y muy usable para el Acto II/III.", "Si", "5,9"),

("Demis Hassabis", "Google DeepMind (via CBS 60 Minutes / TIME)",
 "Radical abundance", "https://www.cbsnews.com/news/artificial-intelligence-google-deepmind-ceo-demis-hassabis-60-minutes-transcript/",
 2025, "La IA elimina la escasez material; el cuello de botella remanente pasa a ser la asignacion y el juicio, no la produccion.",
 "Entrevistas publicas de Hassabis sobre el horizonte de la IA.",
 "That's what I mean by radical abundance. We no longer, in a meaningful way, are in a zero-sum resource-constrained world.",
 "AI-execution", 3, "Refuerza el marco de abundancia de ejecucion vs escasez de juicio.", "Mencion en notas", "9"),

("Irving Janis", "Houghton Mifflin",
 "Victims of Groupthink (libro)", "https://archive.org/details/groupthinkpsycho00jani/",
 1972, "Grupos cohesivos y de alto consenso (Bahia de Cochinos, Pearl Harbor) suprimen la disidencia y producen peores decisiones que individuos solos.",
 "Analisis historico de fiascos de politica exterior de EE.UU.",
 "", "contradiction", 5, "Base clasica del 'gente inteligente, sistema estupido'.", "Si", "21"),

("Ivan D. Steiner", "Academic Press",
 "Group Process and Productivity (libro)", "https://www.scirp.org/reference/referencespapers?referenceid=2875973",
 1972, "Formaliza la 'perdida de proceso': productividad real = productividad potencial menos perdidas de coordinacion y motivacion.",
 "Marco teorico clasico de psicologia de grupos.",
 "Actual productivity = Potential productivity minus Process losses.",
 "contradiction", 4, "Version academica de 'demasiados cocineros'.", "Si", "21"),

("Karen Jehn, Gregory Northcraft, Margaret Neale (y literatura de faultlines)", "Group Decision and Negotiation",
 "Cracks in Diversity Research: The Effects of Diversity Faultlines on Conflict and Performance", "https://link.springer.com/article/10.1023/A:1023325406946",
 2010, "El efecto de la diversidad es curvilineo via 'lineas de fractura': divisiones en subgrupos por variables demograficas correlacionadas aumentan el conflicto.",
 "Revision de literatura sobre faultlines organizacionales.",
 "", "contradiction", 4, "Complica la narrativa simple de 'mas diversidad = mejores decisiones'.", "Si", "22"),

("McKinsey & Company", "McKinsey Quarterly",
 "Three keys to faster, better decisions", "https://www.mckinsey.com/capabilities/people-and-organizational-performance/our-insights/three-keys-to-faster-better-decisions",
 2019, "Solo ~20% de las organizaciones sobresale en toma de decisiones; 61% dice que al menos la mitad del tiempo de decision se usa mal.",
 "Encuesta global; una empresa Fortune 500 tipica pierde ~530,000 dias-gerente/ano (~250M USD) por procesos de decision deficientes.",
 "", "industry-report", 5, "Estadistica ancla del Acto III (el cuello de botella ya existia antes de la IA).", "Si", "11"),

("Eric Bogert, Aaron Schecter, Richard T. Watson", "Science Advances",
 "Generative AI enhances individual creativity but reduces the collective diversity of novel content", "https://www.science.org/doi/10.1126/sciadv.adn5290",
 2024, "Escritores que usan IA generativa producen historias individualmente mas creativas, pero el conjunto de historias resultante es mas similar entre si.",
 "Experimento controlado de escritura asistida por IA generativa.",
 "Individually novel but collectively homogeneous.",
 "contradiction", 5, "Evidencia clave: ejecucion individual mas barata puede reducir la inteligencia colectiva si no se disena para evitarlo.", "Si", "28"),

("Christoph Riedl, David De Cremer", "Collective Intelligence (journal)",
 "AI for Collective Intelligence", "https://journals.sagepub.com/doi/10.1177/26339137251328909",
 2025, "Revisa condiciones bajo las cuales la IA aumenta genuinamente la inteligencia colectiva vs. solo automatiza tareas individuales.",
 "Revision academica reciente sobre IA e inteligencia colectiva.",
 "", "collective-intelligence", 3, "Advierte que ganancias individuales de productividad con IA no se agregan automaticamente en mejores decisiones grupales.", "Mencion en notas", "28,29"),

("Cobertura periodistica de casos de Holacracia (GitHub, Medium, Buffer, Zappos)", "Medium / Nonprofit Quarterly",
 "Holacracy and the mirage of the boss-less workplace / Autopsy of a Failed Holacracy", "https://medium.com/battle-room/holacracy-and-the-mirage-of-the-boss-less-workplace-lessons-from-the-failures-at-github-medium-4355993926d4",
 2020, "Experimentos radicales de descentralizacion/autogestion en empresas conocidas produjeron confusion de roles y agobio, no mejor juicio adaptativo.",
 "Analisis periodistico basado en testimonios de empleados de las empresas mencionadas.",
 "25% reported feeling overwhelmed by the lack of structure.",
 "contradiction", 4, "Contrapeso directo a 'descentralizar siempre es positivo'.", "Si", "34"),

("Gartner", "Gartner Research",
 "4 AI Organization Structure Innovations to Combat Uncertainty / Organization Design insights", "https://www.gartner.com/en/human-resources/insights/organization-design",
 2024, "Para 2026, 20% de las organizaciones usara IA para aplanar la estructura y eliminar mas de la mitad de los puestos de mandos medios actuales.",
 "Investigacion de analistas de Gartner sobre diseno organizacional en la era de IA.",
 "", "industry-report", 3, "Evidencia de cambio estructural anticipado por analistas de industria.", "Mencion en notas", "32"),

("Gartner", "Technology Magazine (cobertura de investigacion Gartner)",
 "How agentic AI is shaping business decision-making", "https://technologymagazine.com/articles/gartner-how-agentic-ai-is-shaping-business-decision-making",
 2024, "Proyecta que los agentes de IA tomaran el 15% de las decisiones cotidianas de trabajo para 2028 (desde ~0%).",
 "Investigacion de analistas sobre IA agentica y decisiones empresariales.",
 "", "industry-report", 3, "Cuantifica el desplazamiento de la ejecucion/decision hacia agentes de IA.", "Mencion en notas", "26,32"),

("McKinsey & Company", "McKinsey Digital",
 "Superagency in the Workplace: Empowering People to Unlock AI's Full Potential", "https://www.mckinsey.com/capabilities/tech-and-ai/our-insights/superagency-in-the-workplace-empowering-people-to-unlock-ais-full-potential-at-work",
 2025, "Encuesta a 3,613 empleados y 238 ejecutivos en 6 paises: los empleados adoptan IA mas rapido de lo que los lideres gobiernan ('agency gap').",
 "Encuesta global de McKinsey Digital.",
 "", "industry-report", 4, "Evidencia de que la brecha de gobierno organizacional, no la tecnologia, es el limite.", "Si", "3,32"),

("McKinsey & Company", "McKinsey People & Organizational Performance",
 "The Agentic Organization: Contours of the Next Paradigm for the AI Era", "https://www.mckinsey.com/capabilities/people-and-organizational-performance/our-insights/the-agentic-organization-contours-of-the-next-paradigm-for-the-ai-era",
 2025, "A medida que la IA absorbe tareas de ejecucion, ganan las estructuras planas con alto intercambio de contexto, no la jerarquia tradicional.",
 "Investigacion de McKinsey; solo ~1% de las empresas se considera en madurez de IA pese a inversion casi universal ('AI theater').",
 "", "industry-report", 4, "Articulacion estructural directa de la division ejecucion/sentido.", "Si", "32"),

("Boston Consulting Group", "BCG",
 "The Widening AI Value Gap / Are You Generating Value from AI?", "https://www.bcg.com/publications/2025/are-you-generating-value-from-ai-the-widening-gap",
 2025, "Solo 22% de las empresas supera la prueba de concepto con IA; el juicio y la toma de decisiones tienen el mayor riesgo de 'des-habilitacion' organizacional.",
 "Investigacion de BCG sobre adopcion de IA a escala.",
 "", "industry-report", 5, "Apoyo directo y cuantificado a la tesis: el juicio, no la ejecucion, es el riesgo/cuello de botella.", "Si", "3,11"),

("Deloitte AI Institute", "Deloitte Insights",
 "State of Generative AI in the Enterprise (Q4) / Decision-making with AI", "https://www.deloitte.com/us/en/insights/topics/talent/human-capital-trends/2026/decision-making-with-ai.html",
 2025, "Las vacantes de mandos medios cayeron mas de 40% (abr 2022-oct 2024); 93% del presupuesto de IA va a tecnologia vs 7% a formar el juicio de las personas.",
 "Encuesta a 3,235 lideres (ago-sept 2025) y reporte trimestral de adopcion empresarial.",
 "No matter how quickly GenAI advances, organizational change only happens so fast.",
 "industry-report", 5, "Estadistica muy fuerte del desequilibrio inversion-en-tecnologia vs inversion-en-juicio.", "Si", "3,29"),

("World Economic Forum", "WEF",
 "The Future of Jobs Report 2025", "https://www.weforum.org/publications/the-future-of-jobs-report-2025/",
 2025, "92 millones de empleos desplazados, 170 millones creados (neto +78M) para 2030; IA/tecnologia de procesamiento de informacion afecta al 86% de las empresas.",
 "Encuesta global a empleadores, informe insignia del WEF.",
 "", "industry-report", 4, "Evidencia macro de que el cambio dejo de ser un evento.", "Si", "3"),

("Suzanne Simard", "Nature / UBC Mother Tree Project",
 "Net transfer of carbon between ectomycorrhizal tree species in the field", "https://www.nature.com/articles/41557",
 1997, "El carbono se mueve entre especies de arboles distintas a traves de redes fungicas micorrizicas compartidas ('wood-wide web').",
 "Estudio de campo pionero, portada de Nature; investigacion posterior del Mother Tree Project sobre arboles 'madre' como nodos centrales.",
 "", "nature-systems", 5, "Base cientifica de la metafora del micelio (Acto IV).", "Si", "14,15"),

("Suzanne Simard", "Knopf",
 "Finding the Mother Tree: Discovering the Wisdom of the Forest (libro)", "https://en.wikipedia.org/wiki/Finding_the_Mother_Tree",
 2021, "Los arboles 'madre', muy conectados, sostienen desproporcionadamente la resiliencia de la red; removerlos daña medible la regeneracion.",
 "Sintesis de decadas de investigacion de campo.",
 "", "nature-systems", 4, "Traduccion organizacional: invertir en los 'nodos hub' de la red humana.", "Si", "15"),

("Deborah M. Gordon", "Princeton University Press / Stanford",
 "Ant Encounters: Interaction Networks and Colony Behavior (libro)", "https://press.princeton.edu/books/paperback/9780691138794/ant-encounters",
 2010, "Ninguna hormiga dirige a la colonia; la asignacion de tareas emerge de la tasa/patron de contactos antenales breves entre hormigas.",
 "Investigacion de campo y modelado de colonias de hormigas como 'computadoras distribuidas'.",
 "", "nature-systems", 4, "Metafora de coordinacion sin autoridad central via senales locales frecuentes.", "Mencion en notas", "16"),

("Thomas D. Seeley", "Princeton University Press / Cornell",
 "Honeybee Democracy (libro)", "https://www.americanscientist.org/article/group-decision-making-in-honey-bee-swarms",
 2010, "Los enjambres de abejas eligen un nuevo nido mediante busqueda distribuida de exploradoras, danzas del meneo proporcionales a la calidad del sitio, y un umbral de quorum (~15-20 exploradoras).",
 "Decadas de experimentos de campo con enjambres de abejas.",
 "", "nature-systems", 5, "Metafora central del Acto IV: decision por quorum, no por unanimidad ni por la voz mas fuerte.", "Si", "16"),

("Iain D. Couzin", "Nature",
 "Effective leadership and decision-making in animal groups on the move", "https://pubmed.ncbi.nlm.nih.gov/15690039/",
 2005, "Una pequena minoria informada (a veces <10%) puede guiar con precision a un grupo grande no informado, gracias a la copia local de senales.",
 "Modelado y experimentos de comportamiento colectivo animal.",
 "", "nature-systems", 4, "Sustenta que pocos individuos alineados pueden guiar a la organizacion sin autoridad formal.", "Si", "17"),

("Iain D. Couzin et al.", "Science",
 "Uninformed individuals promote democratic consensus in animal groups", "https://pubmed.ncbi.nlm.nih.gov/22174256/",
 2011, "Anadir individuos no informados a un grupo dividido entre dos facciones informadas restaura el consenso democratico y diluye a la minoria vocal.",
 "Experimentos de comportamiento colectivo (peces) y modelado.",
 "", "nature-systems", 5, "Hallazgo contraintuitivo y muy potente para la slide de murmuraciones.", "Si", "17"),

("Multiples autores (revision)", "PMC / Annual Review of Immunology",
 "Multiscale information processing in the immune system", "https://pmc.ncbi.nlm.nih.gov/articles/PMC12319014/",
 2024, "El sistema inmune es una red adaptativa multiescala: deteccion barata y generica en todas partes (innata) + amplificacion clonal precisa (adaptativa) + memoria.",
 "Revision cientifica reciente sobre topologia de red del sistema inmune.",
 "", "nature-systems", 4, "Metafora de deteccion distribuida + escalado rapido de respuesta confirmada.", "Si", "18"),

("C.S. Holling", "Annual Review of Ecology and Systematics",
 "Resilience and Stability of Ecological Systems", "https://sciences.ucf.edu/biology/d4lab/wp-content/uploads/sites/23/2024/08/Holling-1973.pdf",
 1973, "La resiliencia (absorber shocks reteniendo estructura/funcion) es distinta de la estabilidad (volver a un equilibrio fijo); introduce el 'ciclo adaptativo'.",
 "Articulo fundacional de la ecologia de sistemas.",
 "", "nature-systems", 5, "Base cientifica del principio: optimizar demasiado la fase de conservacion genera fragilidad.", "Si", "19"),

("Lance Gunderson, C.S. Holling (eds.)", "Island Press",
 "Panarchy: Understanding Transformations in Human and Natural Systems (libro)", "https://islandpress.org/books/panarchy",
 2002, "Los ciclos adaptativos se anidan en escalas (parche-rodal-paisaje); la perturbacion pequena puede ser absorbida o, a veces, escalar y forzar reorganizacion mayor.",
 "Marco teorico de panarquia en ecologia y sistemas socioecologicos.",
 "", "nature-systems", 3, "Sustenta el uso de 'disrupciones controladas' (pilotos, proyectos cancelados) como estrategia.", "Mencion en notas", "19"),

("Atsushi Tero, Toshiyuki Nakagaki et al.", "Science",
 "Rules for Biologically Inspired Adaptive Network Design", "https://www.science.org/doi/10.1126/science.1177894",
 2010, "Un moho mucilaginoso sin cerebro (Physarum polycephalum) diseno una red de transporte con eficiencia comparable a la red ferroviaria de Tokio, mediante poda y refuerzo de flujo.",
 "Experimento con copos de avena dispuestos como ciudades del area de Tokio.",
 "", "nature-systems", 3, "Metafora de bonus/reserva: explorar mucho primero, podar segun el uso real.", "Mencion en notas (bonus, no en slide principal)", "n/a"),

("Toshihiro Nishiguchi, Alexandre Beaudet", "Sloan Management Review",
 "The Toyota Group and the Aisin Fire", "https://www.thecasecentre.org/educators/products/view&&id=6495",
 1998, "Tras el incendio de Aisin en 1997, mas de 200 proveedores se autoorganizaron sin coordinacion central y restauraron la produccion en dias.",
 "Estudio de caso academico del MIT Sloan sobre la recuperacion de la cadena de suministro de Toyota.",
 "", "corporate-story", 5, "Historia empresarial ancla del Acto V/VIII: red de confianza actuando en paralelo, sin plan central.", "Si", "23"),

("Multiples fuentes (TPS)", "Toyota / Lean Blog",
 "Andon (manufacturing)", "https://en.wikipedia.org/wiki/Andon_(manufacturing)",
 2012, "Cualquier trabajador de linea puede detener la produccion al detectar un problema (Jidoka); la autoridad de deteccion se distribuye deliberadamente al borde.",
 "Practica documentada del Sistema de Produccion Toyota.",
 "", "corporate-story", 4, "Ejemplo real de deteccion distribuida diseñada en un sistema de produccion.", "Si", "33"),

("Computer History Museum / Sequoia Capital / CBS News", "Multiples",
 "Nvidia's pivot from gaming GPUs to AI computing", "https://computerhistory.org/profile/jensen-huang/",
 2012, "Nvidia subsidio CUDA sin mercado claro durante 6 anos (desde 2006); cuando AlexNet uso GPUs Nvidia en 2012, la empresa doblo la apuesta una decada antes de ChatGPT.",
 "Historia documentada del pivote estrategico de Nvidia.",
 "", "corporate-story", 4, "Ejemplo de apuesta por una senal debil sensada, no por demanda confirmada.", "Mencion en notas", "n/a"),

("The Hollywood Reporter / Coleman Insights", "Multiples",
 "Netflix's streaming pivot and its own harsh structural choice", "https://www.hollywoodreporter.com/business/business-news/netflix-dvd-by-mail-era-1235402250/",
 2023, "Netflix lanzo streaming en 2007 mientras el DVD aun crecia, y dejo deliberadamente de invitar al equipo de DVD a las reuniones de estrategia para forzar la atencion organizacional.",
 "Declaraciones de Ted Sarandos reportadas en prensa especializada.",
 "", "corporate-story", 4, "Leccion estructural (quien tiene asiento en la mesa) mas que tecnologica.", "Si", "34"),

("Jeremiah Lee (ex Spotify)", "jeremiahlee.com",
 "Spotify's Failed #SquadGoals", "https://www.jeremiahlee.com/posts/failed-squad-goals/",
 2020, "El 'modelo Spotify' de squads autonomos, copiado globalmente, fue documentado por un ex-empleado como abandonado internamente: autonomia sin coordinacion causo duplicacion y luchas de poder.",
 "Analisis interno detallado de un ex gerente de producto de Spotify/Stripe.",
 "", "corporate-story", 5, "Contraejemplo esencial: la descentralizacion no es automaticamente positiva a escala.", "Si", "34"),

("NASA / Smithsonian / IEEE Spectrum", "Multiples",
 "Apollo 13: the CO2 scrubber fix ('square peg in a round hole')", "https://airandspace.si.edu/apollo-missions/apollo-13",
 1970, "Ingenieros en tierra y la tripulacion improvisaron en tiempo real un adaptador de filtro de CO2 con materiales a bordo, bajo una restriccion de tiempo extrema.",
 "Documentacion historica de la mision Apollo 13 y su recuperacion.",
 "", "human-collective-intelligence", 5, "Historia humana ancla: nadie por si solo pudo resolverlo; resolucion paralela y distribuida.", "Si", "23 (alt)"),

("Foldit Contenders Group et al.", "Nature Structural & Molecular Biology",
 "Crystal structure of a monomeric retroviral protease solved by protein folding game players", "https://www.ncbi.nlm.nih.gov/pmc/articles/PMC3211970/",
 2011, "Jugadores del juego Foldit, sin formacion en bioquimica, resolvieron en ~10 dias una estructura proteica que profesionales no habian resuelto en 15 anos.",
 "Publicacion revisada por pares con el grupo de jugadores como coautor acreditado.",
 "", "human-collective-intelligence", 5, "Caso documentado, revisado por pares, de amateurs superando a expertos colectivamente.", "Si", "30"),

("Calcalist/CTech / Dealroom / Lenny's Newsletter", "Multiples",
 "Base44: solo founder builds and sells an AI app-building company for $80M", "https://www.calcalistech.com/ctechnews/article/s1iflnlelx",
 2025, "Maor Shlomo construyo Base44 practicamente solo, alcanzo 1M USD de ARR en 3 semanas y vendio la empresa de 6 meses a Wix por 80M USD, sin equipo ni capital de riesgo.",
 "Cobertura periodistica y entrevista directa al fundador.",
 "", "ai-execution-story", 5, "Ejemplo nitido y reciente de ejecucion individual colapsada por IA.", "Si", "6"),

("Forbes / Fast Company", "Multiples",
 "Klarna's AI assistant: 700 workers' worth of work, then a correction", "https://www.forbes.com/sites/jackkelly/2024/03/04/klarnas-ai-assistant-is-doing-the-job-of-700-workers-company-says/",
 2024, "El asistente de IA de Klarna gestiono 2.3M de chats en su primer mes (equivalente a 700 agentes); en 2025 el CEO admitio haber 'cortado demasiado, demasiado rapido' y reincorporo humanos.",
 "Comunicados de Klarna y entrevista posterior del CEO a Bloomberg.",
 "", "ai-execution-story", 5, "La historia mas clara del keynote: ejecucion resuelta, juicio fallo primero.", "Si", "26,27"),
]

def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Research Database"

    title_font = Font(name="Arial", size=14, bold=True, color="1F2937")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5233")
    body_font = Font(name="Arial", size=9.5)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Bioadaptability / Collective Intelligence / AI — Research Database"
    ws["A1"].font = title_font
    ws.merge_cells("A1:M1")
    ws["A2"] = "Fuente: investigacion realizada via busqueda web en vivo (WebSearch) para el keynote. Todas las URLs fueron localizadas mediante busqueda real; verificar antes de publicacion final si se requiere confirmacion adicional de pagina."
    ws["A2"].font = Font(name="Arial", size=8, italic=True, color="6B7280")
    ws.merge_cells("A2:M2")

    header_row = 4
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for i, row in enumerate(ROWS, start=header_row + 1):
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = body_font
            c.alignment = wrap
            c.border = border
        ws.row_dimensions[i].height = 60

    widths = [26, 22, 40, 34, 7, 42, 42, 40, 18, 9, 40, 16, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:M{header_row + len(ROWS)}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Resumen por categoria"
    ws2["A1"].font = title_font
    cats = {}
    for r in ROWS:
        cats[r[8]] = cats.get(r[8], 0) + 1
    ws2["A3"] = "categoria"
    ws2["B3"] = "num_fuentes"
    ws2["A3"].font = header_font; ws2["A3"].fill = header_fill
    ws2["B3"].font = header_font; ws2["B3"].fill = header_fill
    for i, (k, v) in enumerate(sorted(cats.items(), key=lambda x: -x[1]), start=4):
        ws2.cell(row=i, column=1, value=k).font = body_font
        ws2.cell(row=i, column=2, value=v).font = body_font
    ws2["A" + str(4 + len(cats) + 1)] = f"Total de fuentes: {len(ROWS)}"
    ws2["A" + str(4 + len(cats) + 1)].font = Font(name="Arial", bold=True)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14

    wb.save(OUT)
    print("Saved", OUT, "rows:", len(ROWS))

if __name__ == "__main__":
    build()
